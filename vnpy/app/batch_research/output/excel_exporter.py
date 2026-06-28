"""
output/excel_exporter.py

ExcelExporter  —  把 BatchBacktestResult 列表导出为多 Sheet Excel 文件

Sheet 布局：
  Sheet1 "回测结果"  全部数据行，冻结首行，自动筛选
  Sheet2 "统计摘要"  聚合指标 key/value 表（均值/最大/最小）
  Sheet3 "Top排名"   按夏普比率排序的前 N 名（取当前可见列）

特性：
- 数值列保留原始 float，由 number_format 控制小数位
- 列宽使用 ColumnDefinition.width 转换，不遍历单元格
- 需要 openpyxl（VeighNa Studio 已预装）
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, TYPE_CHECKING

from .exporter import BaseExporter, ExportResult, ExportScope

if TYPE_CHECKING:
    from ..batch_result import BatchBacktestResult
    from ..column_definition import ColumnDefinition
    from ..column_manager import ColumnManager

_FILL_GREEN  = "FFD6F5D6"
_FILL_RED    = "FFFFD6D6"
_FILL_HEADER = "FF1E3A5F"
_FONT_HEADER = "FFFFFFFF"

_FMT_MAP: dict[str, str] = {
    "pct":    "0.00",
    "float1": "0.0",
    "float2": "0.00",
    "float3": "0.000",
    "int":    "0",
    "money":  "#,##0",
    "str":    "@",
}


def _openpyxl_imports():
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
        return openpyxl, PatternFill, Font, Alignment, get_column_letter
    except ImportError as e:
        raise ImportError(
            "openpyxl is required for ExcelExporter. "
            "Install with: pip install openpyxl"
        ) from e


class ExcelExporter(BaseExporter):
    """
    Excel 导出器（多 Sheet）。

    用法::

        exporter = ExcelExporter()
        result = exporter.export(
            bbr_list,
            Path("out.xlsx"),
            column_manager=cm,
            scope=ExportScope.ALL,
        )
        print(result)
    """

    def export(
        self,
        results: list["BatchBacktestResult"],
        filepath: Path | str,
        column_manager: "ColumnManager",
        scope: ExportScope = ExportScope.ALL,
        top_n: int = 20,
    ) -> ExportResult:
        self._top_n           = top_n
        self._column_manager  = column_manager
        return super().export(results, filepath, column_manager, scope)

    def _write(
        self,
        rows: list[dict[str, Any]],
        cols: list["ColumnDefinition"],
        filepath: Path,
    ) -> ExportResult:
        openpyxl, PatternFill, Font, Alignment, get_column_letter = _openpyxl_imports()
        filepath.parent.mkdir(parents=True, exist_ok=True)

        green_fill  = PatternFill(fill_type="solid", fgColor=_FILL_GREEN)
        red_fill    = PatternFill(fill_type="solid", fgColor=_FILL_RED)
        header_fill = PatternFill(fill_type="solid", fgColor=_FILL_HEADER)
        header_font = Font(bold=True, color=_FONT_HEADER)
        center_aln  = Alignment(horizontal="center")

        wb = openpyxl.Workbook()

        self._build_sheet1(wb, rows, cols,
                           green_fill, red_fill, header_fill, header_font, center_aln,
                           get_column_letter)
        self._build_sheet2(wb, rows, cols, header_fill, header_font)
        self._build_sheet3(wb, rows, header_fill, header_font, get_column_letter)

        wb.save(filepath)
        file_size = filepath.stat().st_size

        return ExportResult(
            filepath=filepath,
            rows=len(rows),
            columns=len(cols),
            file_size=file_size,
            success=True,
        )

    def _build_sheet1(self, wb, rows, cols,
                      green_fill, red_fill, header_fill, header_font, center_aln,
                      get_column_letter) -> None:
        ws = wb.active
        ws.title = "回测结果"

        headers = [c.export_header for c in cols]
        keys    = [c.key for c in cols]

        ws.append(headers)
        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = center_aln

        # 列宽：用 ColumnDefinition.width（像素）/ 7 转换为 Excel 字符宽
        for i, col in enumerate(cols):
            letter = get_column_letter(i + 1)
            ws.column_dimensions[letter].width = max(8, col.width / 7)

        return_col_idx = next(
            (i + 1 for i, c in enumerate(cols) if c.key == "total_return"),
            None,
        )

        for row_dict in rows:
            values = []
            for key, col in zip(keys, cols):
                raw = row_dict.get(key)
                if raw is None:
                    values.append("")
                elif col.fmt == "str":
                    values.append(str(raw))
                else:
                    try:
                        values.append(float(raw))
                    except (TypeError, ValueError):
                        values.append(str(raw) if raw is not None else "")
            ws.append(values)

            last_row = ws.max_row
            for i, col in enumerate(cols):
                ws.cell(row=last_row, column=i + 1).number_format = (
                    _FMT_MAP.get(col.fmt, "General")
                )

            # total_return 正绿负红（整行着色）
            if return_col_idx:
                ret_raw = row_dict.get("total_return")
                try:
                    fill = green_fill if float(ret_raw or 0) >= 0 else red_fill
                    for ci in range(1, len(cols) + 1):
                        ws.cell(row=last_row, column=ci).fill = fill
                except (TypeError, ValueError):
                    pass

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def _build_sheet2(self, wb, rows, cols, header_fill, header_font) -> None:
        ws = wb.create_sheet("统计摘要")

        ws.append(["指标", "均值", "最大值", "最小值"])
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        numeric_fmts = {"pct", "float1", "float2", "float3", "int", "money"}
        summary_base = [
            ("股票总数",   len(rows)),
            ("成功数量",   sum(1 for r in rows if r.get("status") == "success")),
            ("失败数量",   sum(1 for r in rows if r.get("status") == "failed")),
            ("跳过数量",   sum(1 for r in rows if r.get("status") == "skipped")),
        ]
        for label, value in summary_base:
            ws.append([label, value, "", ""])

        # 只对 group in (return, risk, trade) 且 default_visible=True 的数值列统计
        stat_groups = {"return", "risk", "trade"}
        stat_cols = [
            col for col in cols
            if col.fmt in numeric_fmts
            and col.group in stat_groups
            and col.default_visible
        ]

        for col in stat_cols:
            vals = []
            for row in rows:
                v = row.get(col.key)
                if v is not None:
                    try:
                        vals.append(float(v))
                    except (TypeError, ValueError):
                        pass
            if not vals:
                continue
            avg  = sum(vals) / len(vals)
            vmax = max(vals)
            vmin = min(vals)
            label = f"均值-{col.export_header}"
            ws.append([label, round(avg, 4), round(vmax, 4), round(vmin, 4)])

        for col_dim in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col_dim), default=8)
            letter = col_dim[0].column_letter
            ws.column_dimensions[letter].width = min(max_len + 2, 40)

    def _build_sheet3(self, wb, rows, header_fill, header_font, get_column_letter) -> None:
        top_n          = getattr(self, "_top_n", 20)
        column_manager = getattr(self, "_column_manager", None)
        ws             = wb.create_sheet(f"Top{top_n}")

        # 取当前可见列中属于 basic/return/risk/trade 组的列
        if column_manager is not None:
            visible = column_manager.get_visible_columns()
            top_cols = [
                c for c in visible
                if c.group in ("basic", "return", "risk", "trade")
            ]
        else:
            # fallback：使用 rows 里所有 key
            top_cols = []

        if not top_cols:
            ws.append(["No columns available"])
            return

        top_headers = [c.export_header for c in top_cols]
        top_keys    = [c.key for c in top_cols]

        ws.append(top_headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for i, col in enumerate(top_cols):
            letter = get_column_letter(i + 1)
            ws.column_dimensions[letter].width = max(8, col.width / 7)

        sorted_rows = sorted(
            [r for r in rows if r.get("status") == "success"],
            key=lambda r: float(r.get("sharpe_ratio") or 0),
            reverse=True,
        )[:top_n]

        for row_dict in sorted_rows:
            values = []
            for key, col in zip(top_keys, top_cols):
                raw = row_dict.get(key)
                if raw is None:
                    values.append("")
                elif col.fmt == "str":
                    values.append(str(raw))
                else:
                    try:
                        values.append(float(raw))
                    except (TypeError, ValueError):
                        values.append(str(raw) if raw is not None else "")
            ws.append(values)

            last_row = ws.max_row
            for i, col in enumerate(top_cols):
                ws.cell(row=last_row, column=i + 1).number_format = (
                    _FMT_MAP.get(col.fmt, "General")
                )

        ws.freeze_panes = "A2"
