"""
ExcelWriter

Writes batch backtest results to an Excel (.xlsx) file.

Sheet layout:
  Sheet 1 "Results"  — one row per symbol, columns follow ORDERED_COLUMNS
  Sheet 2 "Summary"  — aggregate metrics (single column key/value table)
  Sheet 3 "TopN"     — top-N symbols by Sharpe ratio (default 20)

Features:
  - Conditional formatting: green/red background for positive/negative returns
  - Frozen header row + auto-filter on Results sheet
  - Column widths auto-fitted to content
  - Requires openpyxl (pre-installed in VeighNa Studio)
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import TYPE_CHECKING

from ..statistics.analyzer import ORDERED_COLUMNS, StatisticsAnalyzer
from ..statistics.metrics import build_aggregate_summary

if TYPE_CHECKING:
    from ..task import BacktestResult


@dataclass
class WriteResult:
    """Summary of a completed write operation."""
    filepath: Path
    rows_written: int
    sheets: list[str]
    file_size_bytes: int

    def __str__(self) -> str:
        kb = self.file_size_bytes / 1024
        return (
            f"WriteResult({self.filepath.name}: "
            f"{self.rows_written} rows, "
            f"sheets={self.sheets}, "
            f"{kb:.1f} KB)"
        )


# Colour constants (openpyxl ARGB hex)
_GREEN_FILL = "FFD6F5D6"   # light green
_RED_FILL   = "FFFFD6D6"   # light red
_HEADER_FILL = "FF4472C4"  # blue header
_HEADER_FONT = "FFFFFFFF"  # white text


def _make_fills():
    from openpyxl.styles import PatternFill
    return (
        PatternFill(fill_type="solid", fgColor=_GREEN_FILL),
        PatternFill(fill_type="solid", fgColor=_RED_FILL),
        PatternFill(fill_type="solid", fgColor=_HEADER_FILL),
    )


def _make_header_font():
    from openpyxl.styles import Font
    return Font(bold=True, color=_HEADER_FONT)


def _autofit_columns(ws, max_width: int = 30) -> None:
    """Set column widths based on maximum content length."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


class ExcelWriter:
    """
    Writes a list of BacktestResult objects to a multi-sheet Excel file.

    Usage::

        writer = ExcelWriter()
        wr = writer.write(results, Path("output/batch_result.xlsx"))
        print(wr)
    """

    def write(
        self,
        results: list["BacktestResult"],
        filepath: Path | str,
        enrich: bool = True,
        top_n: int = 20,
    ) -> WriteResult:
        """
        Write results to an Excel .xlsx file.

        :param results:   List of BacktestResult objects.
        :param filepath:  Output path (created if needed; must end in .xlsx).
        :param enrich:    Run StatisticsAnalyzer.enrich() before writing.
        :param top_n:     Number of top-Sharpe symbols shown in TopN sheet.
        :return:          WriteResult with counts and file size.
        """
        try:
            import openpyxl
            from openpyxl.styles import Alignment
        except ImportError as e:
            raise ImportError(
                "openpyxl is required for ExcelWriter. "
                "Install with: pip install openpyxl"
            ) from e

        filepath = Path(filepath)
        filepath.parent.mkdir(parents=True, exist_ok=True)

        if not results:
            wb = openpyxl.Workbook()
            wb.save(filepath)
            return WriteResult(filepath=filepath, rows_written=0,
                               sheets=[], file_size_bytes=0)

        analyzer = StatisticsAnalyzer()
        if enrich:
            analyzer.enrich(results)

        rows = [r.to_flat_dict() for r in results]
        all_keys = list(
            dict.fromkeys(
                [c for c in ORDERED_COLUMNS if c in rows[0]]
                + [k for k in rows[0] if k not in ORDERED_COLUMNS]
            )
        )

        wb = openpyxl.Workbook()
        green_fill, red_fill, header_fill = _make_fills()
        header_font = _make_header_font()

        # ── Sheet 1: Results ──────────────────────────────────────── #
        ws_results = wb.active
        ws_results.title = "Results"

        # Header row
        ws_results.append(all_keys)
        for cell in ws_results[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data rows with conditional fill on total_return
        return_col_idx = (all_keys.index("total_return") + 1
                          if "total_return" in all_keys else None)

        for row_dict in rows:
            row_values = [row_dict.get(k, "") for k in all_keys]
            ws_results.append(row_values)

            if return_col_idx is not None:
                ret_val = row_dict.get("total_return", 0)
                try:
                    fill = green_fill if float(ret_val) >= 0 else red_fill
                except (TypeError, ValueError):
                    fill = None
                if fill:
                    last_row = ws_results.max_row
                    for col_idx in range(1, len(all_keys) + 1):
                        ws_results.cell(row=last_row, column=col_idx).fill = fill

        # Freeze header + auto-filter
        ws_results.freeze_panes = "A2"
        ws_results.auto_filter.ref = ws_results.dimensions
        _autofit_columns(ws_results)

        # ── Sheet 2: Summary ──────────────────────────────────────── #
        ws_summary = wb.create_sheet("Summary")
        summary = build_aggregate_summary(results)

        ws_summary.append(["Metric", "Value"])
        for cell in ws_summary[1]:
            cell.fill = header_fill
            cell.font = header_font

        label_map = {
            "agg_total_symbols":     "Total Symbols",
            "agg_success_symbols":   "Successful",
            "agg_failed_symbols":    "Failed",
            "agg_skipped_symbols":   "Skipped",
            "agg_avg_total_return":  "Avg Total Return (%)",
            "agg_avg_annual_return": "Avg Annual Return (%)",
            "agg_avg_sharpe":        "Avg Sharpe Ratio",
            "agg_avg_max_ddpercent": "Avg Max Drawdown (%)",
            "agg_avg_calmar":        "Avg Calmar Ratio",
            "agg_win_rate":          "Win Rate (%)",
            "agg_profit_loss_ratio": "Profit/Loss Ratio",
            "agg_total_trades":      "Total Trades",
            "agg_avg_trades":        "Avg Trades / Symbol",
        }
        for key, label in label_map.items():
            ws_summary.append([label, summary.get(key, "")])

        _autofit_columns(ws_summary, max_width=40)

        # ── Sheet 3: TopN ─────────────────────────────────────────── #
        ws_topn = wb.create_sheet(f"Top{top_n}")
        top_results = analyzer.top_n(results, n=top_n, by="sharpe_ratio")

        top_cols = [
            "vt_symbol", "total_return", "annual_return",
            "sharpe_ratio", "calmar_ratio", "max_ddpercent",
            "max_drawdown_duration", "total_trade_count",
            "return_drawdown_ratio", "ewm_sharpe",
        ]
        top_cols_exist = [c for c in top_cols if any(c in r.to_flat_dict() for r in top_results)]

        ws_topn.append(top_cols_exist)
        for cell in ws_topn[1]:
            cell.fill = header_fill
            cell.font = header_font

        for r in top_results:
            flat = r.to_flat_dict()
            ws_topn.append([flat.get(c, "") for c in top_cols_exist])

        ws_topn.freeze_panes = "A2"
        _autofit_columns(ws_topn)

        wb.save(filepath)

        file_size = filepath.stat().st_size
        sheets = [ws_results.title, ws_summary.title, ws_topn.title]
        return WriteResult(
            filepath=filepath,
            rows_written=len(rows),
            sheets=sheets,
            file_size_bytes=file_size,
        )
