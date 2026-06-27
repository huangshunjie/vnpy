"""Phase 7 test Part 2: ExcelWriter + Integration"""

import copy
import random
import tempfile
from datetime import datetime, timedelta
from pathlib import Path

from vnpy.trader.constant import Exchange, Interval
from vnpy.trader.object import BarData
from vnpy_ctastrategy.strategies.atr_rsi_strategy import AtrRsiStrategy

from vnpy.app.batch_research.task import BacktestResult, TaskStatus
from vnpy.app.batch_research.batch_engine import BatchBacktestingEngine
from vnpy.app.batch_research.output.excel_writer import ExcelWriter
from vnpy.app.batch_research.output.csv_writer import CSVWriter


# ================================================================
# Helpers
# ================================================================

def make_mock_result(
    vt_symbol: str,
    total_return: float = 10.0,
    annual_return: float = 5.0,
    sharpe: float = 1.2,
    max_ddpercent: float = -8.0,
    total_trade_count: int = 50,
    status: TaskStatus = TaskStatus.SUCCESS,
) -> BacktestResult:
    result = BacktestResult(
        vt_symbol=vt_symbol,
        task_id=f"t_{vt_symbol}",
        strategy_name="TestStrategy",
        status=status,
    )
    if status == TaskStatus.SUCCESS:
        result.statistics = {
            "start_date": "2020-01-02", "end_date": "2021-06-30",
            "total_days": 365, "profit_days": 200, "loss_days": 165,
            "capital": 1_000_000,
            "end_balance": 1_000_000 * (1 + total_return / 100),
            "total_return": total_return, "annual_return": annual_return,
            "daily_return": annual_return / 240, "return_std": 1.5,
            "max_drawdown": max_ddpercent * 10000,
            "max_ddpercent": max_ddpercent, "max_drawdown_duration": 30,
            "sharpe_ratio": sharpe, "ewm_sharpe": sharpe * 0.9,
            "return_drawdown_ratio": abs(total_return / max_ddpercent) if max_ddpercent else 0,
            "rgr_ratio": 1.5,
            "total_net_pnl": total_return * 10000,
            "daily_net_pnl": total_return * 10000 / 365,
            "total_commission": 5000, "daily_commission": 5000 / 365,
            "total_slippage": 2000,  "daily_slippage": 2000 / 365,
            "total_turnover": 5_000_000, "daily_turnover": 5_000_000 / 365,
            "total_trade_count": total_trade_count,
            "daily_trade_count": total_trade_count / 365,
        }
    if status == TaskStatus.FAILED:
        result.error_msg = "mock error"
    return result


MOCK_RESULTS = [
    make_mock_result("000001.SZSE", total_return=15.0, sharpe=1.8,  max_ddpercent=-8.0),
    make_mock_result("600519.SSE",  total_return=25.0, sharpe=2.5,  max_ddpercent=-5.0),
    make_mock_result("300750.SZSE", total_return=-3.0, sharpe=-0.5, max_ddpercent=-12.0),
    make_mock_result("000858.SZSE", total_return=8.0,  sharpe=1.1,  max_ddpercent=-6.5),
    make_mock_result("600036.SSE",  total_return=-5.0, sharpe=-0.8, max_ddpercent=-15.0),
    make_mock_result("688001.SSE",  status=TaskStatus.FAILED),
    make_mock_result("301500.SZSE", status=TaskStatus.SKIPPED),
]


def make_bars(symbol: str = "000001", n: int = 400) -> list[BarData]:
    rng = random.Random(hash(symbol) & 0xFFFF)
    bars: list[BarData] = []
    price = 10.0
    dt = datetime(2020, 1, 2)
    for _ in range(n):
        while dt.weekday() >= 5:
            dt += timedelta(days=1)
        chg = rng.gauss(0, 0.015)
        o = price
        c = round(max(0.1, price * (1 + chg)), 2)
        h = round(max(o, c) * (1 + abs(rng.gauss(0, 0.004))), 2)
        l = round(min(o, c) * (1 - abs(rng.gauss(0, 0.004))), 2)
        bars.append(BarData(
            gateway_name="CSV", symbol=symbol,
            exchange=Exchange.SZSE, datetime=dt,
            interval=Interval.DAILY,
            open_price=o, high_price=h, low_price=l, close_price=c,
            volume=float(rng.randint(500_000, 3_000_000)),
        ))
        price = c
        dt += timedelta(days=1)
    return bars


# ================================================================
# ExcelWriter tests
# ================================================================

def test_excel_writer_basic():
    import openpyxl
    results = copy.deepcopy(MOCK_RESULTS)
    with tempfile.TemporaryDirectory() as tmpdir:
        fp = Path(tmpdir) / "results.xlsx"
        wr = ExcelWriter().write(results, fp)

        assert fp.exists()
        assert wr.rows_written == len(results)
        assert wr.file_size_bytes > 0
        assert "Results" in wr.sheets
        assert "Summary" in wr.sheets

        wb = openpyxl.load_workbook(fp)
        ws = wb["Results"]
        header = [cell.value for cell in ws[1]]
        assert "vt_symbol" in header
        assert "sharpe_ratio" in header
        assert ws.max_row == len(results) + 1   # header + data rows

    print(f"PASS  ExcelWriter basic  {wr}")


def test_excel_writer_summary_sheet():
    import openpyxl
    results = copy.deepcopy(MOCK_RESULTS)
    with tempfile.TemporaryDirectory() as tmpdir:
        fp = Path(tmpdir) / "results.xlsx"
        ExcelWriter().write(results, fp)
        wb = openpyxl.load_workbook(fp)
        ws = wb["Summary"]
        labels = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert "Total Symbols" in labels
        assert "Win Rate (%)" in labels
        assert "Avg Sharpe Ratio" in labels
    print("PASS  ExcelWriter Summary sheet")


def test_excel_writer_topn_sheet():
    import openpyxl
    results = copy.deepcopy(MOCK_RESULTS)
    with tempfile.TemporaryDirectory() as tmpdir:
        fp = Path(tmpdir) / "results.xlsx"
        ExcelWriter().write(results, fp, top_n=3)
        wb = openpyxl.load_workbook(fp)
        topn_sheet = [s for s in wb.sheetnames if s.startswith("Top")][0]
        ws = wb[topn_sheet]
        data_rows = ws.max_row - 1
        assert data_rows <= 3
        header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        assert "vt_symbol" in header
    print(f"PASS  ExcelWriter TopN sheet  ({data_rows} rows)")


def test_excel_writer_frozen_header():
    """Results sheet must have frozen pane at A2."""
    import openpyxl
    results = copy.deepcopy(MOCK_RESULTS[:3])
    with tempfile.TemporaryDirectory() as tmpdir:
        fp = Path(tmpdir) / "frozen.xlsx"
        ExcelWriter().write(results, fp)
        wb = openpyxl.load_workbook(fp)
        ws = wb["Results"]
        assert ws.freeze_panes == "A2"
    print("PASS  ExcelWriter frozen header row")


def test_excel_writer_autofilter():
    """Results sheet must have auto-filter enabled."""
    import openpyxl
    results = copy.deepcopy(MOCK_RESULTS[:3])
    with tempfile.TemporaryDirectory() as tmpdir:
        fp = Path(tmpdir) / "autofilter.xlsx"
        ExcelWriter().write(results, fp)
        wb = openpyxl.load_workbook(fp)
        ws = wb["Results"]
        assert ws.auto_filter.ref is not None
    print("PASS  ExcelWriter auto-filter enabled")


def test_excel_writer_calmar_enriched():
    """calmar_ratio column must appear in Results sheet (added by enrich=True)."""
    import openpyxl
    results = copy.deepcopy(MOCK_RESULTS[:3])
    with tempfile.TemporaryDirectory() as tmpdir:
        fp = Path(tmpdir) / "enriched.xlsx"
        ExcelWriter().write(results, fp, enrich=True)
        wb = openpyxl.load_workbook(fp)
        header = [cell.value for cell in wb["Results"][1]]
        assert "calmar_ratio" in header
    print("PASS  ExcelWriter calmar_ratio column present")


def test_excel_writer_empty_results():
    import openpyxl
    with tempfile.TemporaryDirectory() as tmpdir:
        fp = Path(tmpdir) / "empty.xlsx"
        wr = ExcelWriter().write([], fp)
        assert wr.rows_written == 0 and fp.exists()
    print("PASS  ExcelWriter empty results")


def test_excel_writer_creates_parent_dir():
    results = copy.deepcopy(MOCK_RESULTS[:2])
    with tempfile.TemporaryDirectory() as tmpdir:
        fp = Path(tmpdir) / "sub" / "nested" / "out.xlsx"
        ExcelWriter().write(results, fp)
        assert fp.exists()
    print("PASS  ExcelWriter creates parent directory")


# ================================================================
# Integration: BatchBacktestingEngine -> export
# ================================================================

def _build_engine(syms: list[str]) -> BatchBacktestingEngine:
    engine = BatchBacktestingEngine()
    engine.set_parameters(
        strategy_class=AtrRsiStrategy,
        start=datetime(2020, 1, 1),
        end=datetime(2021, 6, 30),
        capital=1_000_000,
        rate=1e-4,
        slippage=0.02,
        size=1.0,
        pricetick=0.01,
        strategy_setting={"atr_length": 22, "atr_ma_length": 10},
    )
    engine.set_stock_pool(syms)
    for vt in syms:
        engine.set_bars(vt, make_bars(vt.split(".")[0]))
    engine.run_backtesting(show_progress=False)
    return engine


def test_integration_csv_export():
    import csv as _csv
    syms = ["000001.SZSE", "600519.SSE", "300750.SZSE"]
    engine = _build_engine(syms)

    with tempfile.TemporaryDirectory() as tmpdir:
        fp = Path(tmpdir) / "output.csv"
        engine.export_to_csv(fp)

        assert fp.exists()
        with open(fp, encoding="utf-8-sig") as f:
            rows = list(_csv.DictReader(f))
        data_rows = [r for r in rows if r.get("vt_symbol") != "__SUMMARY__"]
        assert len(data_rows) == 3
        assert "sharpe_ratio" in rows[0]

    print(f"PASS  Integration CSV export  ({len(data_rows)} symbol rows)")


def test_integration_excel_export():
    import openpyxl
    syms = ["000001.SZSE", "600519.SSE", "300750.SZSE"]
    engine = _build_engine(syms)

    with tempfile.TemporaryDirectory() as tmpdir:
        fp = Path(tmpdir) / "output.xlsx"
        engine.export_to_excel(fp)

        assert fp.exists()
        wb = openpyxl.load_workbook(fp)
        assert "Results" in wb.sheetnames
        assert "Summary" in wb.sheetnames
        ws = wb["Results"]
        assert ws.max_row == len(syms) + 1

    print(f"PASS  Integration Excel export  (sheets={wb.sheetnames})")


def test_integration_analyzer_with_real_results():
    """Run full pipeline and apply StatisticsAnalyzer on real results."""
    from vnpy.app.batch_research.statistics.analyzer import StatisticsAnalyzer
    syms = ["000001.SZSE", "600519.SSE", "300750.SZSE",
            "000858.SZSE", "600036.SSE"]
    engine = _build_engine(syms)

    analyzer = StatisticsAnalyzer()
    results = engine.results

    # Enrich adds calmar_ratio and profit_factor
    analyzer.enrich(results)
    for r in [r for r in results if r.statistics]:
        assert "calmar_ratio" in r.statistics
        assert "profit_factor" in r.statistics

    # Summary
    summary = analyzer.summarize(results)
    assert summary["agg_total_symbols"] == 5
    assert 0 <= summary["agg_win_rate"] <= 100

    # DataFrame
    df = analyzer.to_dataframe(results)
    assert len(df) == len([r for r in results if r.statistics])
    assert "calmar_ratio" in df.columns

    # print_summary smoke test
    analyzer.print_summary(results, top_n=3)

    print(f"PASS  Integration analyzer  "
          f"win={summary['agg_win_rate']:.1f}%  "
          f"avg_sharpe={summary['agg_avg_sharpe']:.2f}  "
          f"df_shape={df.shape}")


# ================================================================
# Entry point
# ================================================================

if __name__ == "__main__":
    print("=" * 65)
    print("Phase 7 Test Part 2: ExcelWriter + Integration")
    print("=" * 65)

    print("\n--- ExcelWriter ---")
    test_excel_writer_basic()
    test_excel_writer_summary_sheet()
    test_excel_writer_topn_sheet()
    test_excel_writer_frozen_header()
    test_excel_writer_autofilter()
    test_excel_writer_calmar_enriched()
    test_excel_writer_empty_results()
    test_excel_writer_creates_parent_dir()

    print("\n--- Integration ---")
    test_integration_csv_export()
    test_integration_excel_export()
    test_integration_analyzer_with_real_results()

    print()
    print("=" * 65)
    print("Phase 7 Part 2 ALL TESTS PASSED")
    print("=" * 65)
